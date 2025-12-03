import os
import sys
import subprocess
import re
import psutil
import chardet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side








####### libs check/install ######################################################################
required_libs = ["psutil", "openpyxl", "chardet"]
for lib in required_libs:
    try:
        __import__(lib)
    except ImportError:
        print(f"[INFO] Installing missing library: {lib}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
####### libs install ############################################################################








########### prep #################################################################################
report_path = r"C:\praktinis assignment\report\report.txt"
report_folder = os.path.dirname(report_path)
if not os.path.exists(report_path):
    print(f"[ERROR] report.txt not found: {report_path}")
    sys.exit(1)
with open(report_path, "rb") as f:
    raw = f.read()
    enc = chardet.detect(raw)["encoding"]
print(f"[INFO] Encoding detected: {enc}")
content = raw.decode(enc, errors="replace")
content = content.replace("\x00", "").replace("\ufeff", "")
lines = [line.strip() for line in content.splitlines() if line.strip()]
print(f"[DEBUG] Total non-empty lines: {len(lines)}")
try:
    reg_success = lines[0].split(": ")[1].strip()      
    login_used = lines[1].split(": ")[1].strip()        
    account_created = lines[2].split(": ")[1].strip()    
    loan_accepted = lines[3].split(": ")[1].strip()      
    missing_info = lines[4].split(": ")[1].strip()        
    missing_initial = lines[5].split(": ")[1].strip()     
except IndexError:
    print("[ERROR] Could not extract 6 boolean values.")
    sys.exit(1)
print(f"[DEBUG] Booleans: {reg_success}, {login_used}, {account_created}, {loan_accepted}, {missing_info}, {missing_initial}")
try:
    result_js = lines[6].split(": ")[1].strip()
    print(f"[DEBUG] Account number: {result_js}")
except IndexError:
    result_js = ""
    print("[WARN] Account number not found")
try:
    loan_reason = lines[7].split(": ")[1].strip()
    print(f"[DEBUG] Loan denial reason: {loan_reason}")
except IndexError:
    loan_reason = ""
    print("[WARN] Loan denial reason not found")
try:
    internal_error = lines[8].split(": ")[1].strip()
    print(f"[DEBUG] Internal error: {internal_error}")
except IndexError:
    internal_error = "false"
    print("[WARN] Internal error flag not found")
csv_line_index = 9
try:
    user_line = lines[csv_line_index]
except IndexError:
    print("[ERROR] Could not find user data row.")
    sys.exit(1)


fields = [x.strip() for x in user_line.split(",")]
(
    first_name, last_name, address, city, state,
    zip_code, phone, ssn, username, password,
    account_type, initial_deposit, dob, debit_card, cvv
) = fields[:15]
output_xlsx = os.path.join(report_folder, f"{username}.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "Customer Report"
bold = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
ws["B2"].value = "Actions performed Report"
ws["B2"].font = bold

########### prep #################################################################################









################# Internal Error #########################################################################
if internal_error == "true":
    error_report_path = os.path.join(report_folder, f"{username}_report.txt")
    try:
        import shutil
        shutil.copy2(report_path, error_report_path)
        print(f"[INFO] Error report saved: {error_report_path}")
    except Exception as e:
        print(f"[WARN] Could not create error report: {e}")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].border = Border() 
    full_message = (
        "Internal Error occurred during account processing.\n"
        "Please contact support for assistance.\n"
        "Error details have been logged for review.\n"
        "Thank you for your patience."
    )
    ws["B3"].value = full_message
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["B3"].border = Border() 
    ws["C3"].value = ""
    ws["C3"].border = Border()
    ws.row_dimensions[3].height = 70
    max_line_len = max(len(line) for line in full_message.split("\n"))
    ws.column_dimensions["B"].width = max_line_len + 4
    goto_render_right_block = True
################# Internal Error #########################################################################










################# missing info ##########################################################################
elif missing_info == "true":
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].border = Border() 
    full_message = (
        "Missing information, cannot Register Account.\n"
        "Please check if any information is missing from your data.\n"
        "If nothing is missing, please contact support.\n"
        "Thank you."
    )
    ws["B3"].value = full_message
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["B3"].border = Border()   
    ws["C3"].value = ""
    ws["C3"].border = Border()
    ws.row_dimensions[3].height = 70
    max_line_len = max(len(line) for line in full_message.split("\n"))
    ws.column_dimensions["B"].width = max_line_len + 4
    goto_render_right_block = True
################# missing info ##########################################################################





################# login/registration method #############################################################
else:
    goto_render_right_block = False
if reg_success == "true" and login_used == "false":
    login_method = "Registration"
elif reg_success == "false" and login_used == "true":
    login_method = "LogIn"
elif reg_success == "true" and login_used == "true":
    login_method = "Registration"
else:
    login_method = "Unknown"
def tf(x):
    return "Yes" if x == "true" else "No" if x == "false" else x
left_labels = [
    "Logged in to Account through method (Registration/LogIn):",
    f"{account_type} Account Opened?"
]
left_values = [
    login_method,
    tf(account_created)
]
################# login/registration method #############################################################





################# Opened Account's Number  ##############################################################
if result_js and result_js not in ["", "UNKNOWN"] and result_js.isdigit():
    left_labels.append("Opened Account's number:")
    left_values.append(int(result_js) if result_js.isdigit() else result_js)
################# Opened Account's Number  ##############################################################




################ Loan request status ####################################################################
loan_word = "Accepted" 
if loan_accepted == "true":
    loan_word = "Accepted" 
if loan_accepted == "false":
    loan_word = "Denied"
if missing_initial == "true":
    left_labels.append("Loan request Accepted/Denied?")
    left_values.append("MISSING_DEPOSIT_PLACEHOLDER")
else:
    left_labels.append("Loan request Accepted/Denied?")
    left_values.append(loan_word)
################ Loan request status ####################################################################




################ Requested loan amount ##################################################################
    left_labels.append("Requested Loan amount:")
    if initial_deposit is not None:
        left_values.append("10000 EUR")
    else:
        left_values.append("Cannot request a loan, missing initial deposit information")
################ Requested loan amount ##################################################################




################ down payment info ######################################################################
    left_labels.append("Down payment amount:")
    if initial_deposit and initial_deposit.strip():
        try:
            deposit_amount = float(initial_deposit)
            down_payment = deposit_amount * 0.20
            left_values.append(f"{down_payment:.2f} EUR")
        except ValueError:
            left_values.append("Invalid initial deposit")
    else:
        left_values.append("Missing initial deposit information")
################ down payment info ######################################################################





############### denial reason ##########################################################################
    if loan_accepted == "false":
        left_labels.append("Reason of Denied Loan:")
        if initial_deposit is None:
            left_values.append("Cannot request a loan, missing initial deposit information")
        else:
            left_values.append(loan_reason if loan_reason and loan_reason != "UKNOWN" else "No reason provided")
############### denial reason ##########################################################################





############## missing initial deposit ################################################################

if not goto_render_right_block:
    row = 3
    for label, value in zip(left_labels, left_values):
        if value == "MISSING_DEPOSIT_PLACEHOLDER":
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"B{row}"].value = label
            ws[f"C{row}"].value = (
                "Missing Initial Deposit information, \n"
                "cannot request a loan. \n"
                "Please update your given information. \n"
                "If no information is missing,\n"
                "Please contact support."
            )
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"C{row}"].font = Font(bold=True, italic=True)
            ws[f"B{row}"].border = thin_border
            ws[f"C{row}"].border = thin_border
            ws.row_dimensions[row].height = 90
            ws.column_dimensions["C"].width = 30
            row += 1
            continue
        ws[f"B{row}"] = label
        ws[f"C{row}"] = value
        ws[f"B{row}"].border = thin_border
        ws[f"C{row}"].border = thin_border
        ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center")
        row += 1
ws["B2"].border = thin_border
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

############## missing initial deposit ################################################################





############# Customer Information ######################################################################
ws["E2"] = "Customer Information"
ws["E2"].font = bold
ws["E2"].border = thin_border
ws["F2"].border = thin_border


info_fields = [
    ("FirstName", first_name),
    ("LastName", last_name),
    ("Address", address),
    ("City", city),
    ("State", state),
    ("ZipCode", int(zip_code) if zip_code.isdigit() else zip_code), 
    ("Phone", int(phone) if phone.isdigit() else phone),  
    ("SSN", ssn),
    ("Username", username),
    ("Password", password),
    ("AccountType", account_type),
    ("InitialDeposit", float(initial_deposit) if initial_deposit and initial_deposit.replace('.','').isdigit() else initial_deposit),  # ← Convert
    ("DOB", dob),
    ("DebitCard", debit_card),
    ("CVV", int(cvv) if cvv.isdigit() else cvv), 
]

row = 3
for label, value in info_fields:
    ws[f"E{row}"] = label
    ws[f"F{row}"] = value

    ws[f"E{row}"].border = thin_border
    ws[f"F{row}"].border = thin_border
    ws[f"F{row}"].alignment = Alignment(horizontal="left", vertical="center")
    row += 1
############# Customer Information ######################################################################








##############ZIP CODE NOte ###############################################################################
zip_invalid = zip_code.isalpha()
if zip_invalid:
    ws["B17"].value = (
        "Note: Please check if your Zip Code\n"
        "information is valid and update it.\n"
        "If its valid, please ignore this message.\n"
        "Thank you."
    )

    ws["B17"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["B17"].font = Font(bold=True)
    ws["B17"].border = Border()  
    ws.row_dimensions[17].height = 85
    ws.column_dimensions["B"].width = max(len(line) for line in ws["B17"].value.split("\n")) + 4
##############ZIP CODE NOte ###############################################################################









############## Additional cosmetics / save )) #############################################################
if missing_info != "true" and internal_error != "true":
    ws.column_dimensions["B"].width = 32
max_c_width = 15  
for row in ws.iter_rows(min_col=3, max_col=3):  
    for cell in row:
        if cell.value:
            cell_length = len(str(cell.value))
            if "\n" in str(cell.value):
                cell_length = max(len(line) for line in str(cell.value).split("\n"))
            if cell_length > max_c_width:
                max_c_width = cell_length
ws.column_dimensions["C"].width = max_c_width + 2
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 25
wb.save(output_xlsx)
print(f"[OK] Excel report created: {output_xlsx}")
############## Additional cosmetics / save )) #############################################################